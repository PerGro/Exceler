from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter

"""
本脚本主要针对检索str内容并为其添加权重，最后计算和并写入excel中
本脚本添加额外依赖为openpyxl
本脚本测试运行环境为Python 3.7
作者：PerGro
第一次写脚本，回头看看简直是龙飞凤舞
今后还需更加深入学习才是
"""


class Excel:
    def __init__(self, path, sheetname, excelname=None):
        self._path = path
        self._sheet_name = sheetname
        self._excel_name = excelname

    """
    set_name方法用来设定保存表格的名字
    """

    def set_name(self, name):
        self._excel_name = name

    """
    set_name_auto方法将自动把源文件当做保存文件名来使用
    """

    @property
    def set_name_auto(self):
        names = self._path[::-1]
        names = names.split('\\')
        names = names[0]
        self._excel_name = names[::-1]

    """
    获取保存excel的名称
    """

    @property
    def get_name(self):
        if self._excel_name:
            print(self._excel_name)
            return self._excel_name
        else:
            raise Exception('初始化失败，请调用set_name_auto或set_name方法')

    """
    获取某列所有不同的数据，顺序以前后出现顺序为准
    number属性为要读取的单元格格数，不填则在读到第一个空单元格为止
    在这以英文逗号为分割，具体情况可再做修改
    """

    def cat_columns(self, rows, columns, numbers=None):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        retlist = []
        values = ws.cell(row=rows, column=columns).value
        if numbers is not None:
            for i in range(numbers):
                values = ws.cell(row=rows, column=columns).value
                value_fixed = values.split(',')
                for n in range(len(value_fixed)):
                    if value_fixed[n] not in retlist:
                        retlist.append(value_fixed[n])
                rows += 1
        else:
            while values is not None:
                value_fixed = values.split(',')
                values = ws.cell(row=rows, column=columns).value
                for n in range(len(value_fixed)):
                    if value_fixed[n] not in retlist:
                        retlist.append(value_fixed[n])
                rows += 1
        return retlist

    """
    与上个方法相同，只不过为读取一行单元格内容
    number属性为要读取的单元格格数，不填则在读到第一个空单元格为止
    """

    def cat_row(self, rows, columns, numbers=None):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        retlist = []
        values = ws.cell(row=rows, column=columns).value
        if numbers is not None:
            for i in range(numbers):
                values = ws.cell(row=rows, column=columns).value
                value_fixed = values.split(',')
                for n in range(len(value_fixed)):
                    if value_fixed[n] not in retlist:
                        retlist.append(value_fixed[n])
                columns += 1
        else:
            while values is not None:
                value_fixed = values.split(',')
                values = ws.cell(row=rows, column=columns).value
                for n in range(len(value_fixed)):
                    if value_fixed[n] not in retlist:
                        retlist.append(value_fixed[n])
                columns += 1
        return retlist

    """
    用来设置权重，在设置时要注意顺序与读取内容顺序一致
    可使用print_info方法来确定读取内容顺序
    """

    def set_data_weight(self, data_list, weight_list):
        mp = dict(zip(data_list, weight_list))
        return mp

    """
    打印读取内容
    """

    def print_info(self, data_list):
        print('目前的加权项目有：')
        for n in range(len(data_list)):
            print(data_list[n])

    """
    打印权重与读取内容，以检查顺序
    """

    def print_weight_info(self, data_list, weight_list):
        return data_list, weight_list

    """
    简单将数据添加到某个单元格中
    """

    def save_to_cell_simple(self, rows, columns, data):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        ws.cell(row=rows, column=columns).value = data
        wb.save(self._excel_name)
        wb.close()

    """
    将简单数据（list）写入某一列中
    numbers为需要写入单元格数量，默认为读取内容遇到的第一个空单元格前同行某列
    rows默认为从第一行开始读入
    """

    def save_to_cell_columns(self, columns, data_list, numbers=None,rows=1):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        cells = ws.cell(row=rows, column=columns).value
        if numbers:
            for n in range(numbers):
                ws.cell(row=rows, column=columns).value = data_list[n]
                rows += 1
        else:
            n = 0
            while cells:
                ws.cell(row=rows, column=columns).value = data_list[n]
                rows += 1
        wb.save(self._excel_name)
        ws.close()

    """
    将简单数据（list）写入某一行中
    numbers为需要写入单元格数量，默认为读取内容遇到的第一个空单元格前同列某行
    columns默认为从第一行开始读入
    """

    def save_to_cell_rows(self, rows, data_list, numbers=None, columns=1):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        cells = ws.cell(row=rows, column=columns).value
        if numbers:
            for n in range(numbers):
                ws.cell(row=rows, column=columns).value = data_list[n]
                columns += 1
        else:
            n = 0
            while cells:
                ws.cell(row=rows, column=columns).value = data_list[n]
                columns += 1
        wb.save(self._excel_name)
        ws.close()

    """
    将加权后数据（dice）写入某一列中
    numbers为需要写入单元格数量，默认为读取内容遇到的第一个空单元格前同行某列
    rows默认为从第一行开始读入
    """

    def count_weight_sum_columns(self, data_weight_dice, fromcolumns, tocolumns, numbers=None, rows=1):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        datalist = list(data_weight_dice.keys())
        weightlist = list(data_weight_dice.values())
        fromcells = ws.cell(row=rows, column=fromcolumns)
        tocells = ws.cell(row=rows, column=tocolumns)
        if numbers:
            for n in range(numbers):
                sumer = 0
                print('fromcells is ', fromcells.value)
                for i in range(len(datalist)):
                    if datalist[i] in fromcells.value:
                        print(fromcells.value)
                        sumer += weightlist[i]
                tocells.value = sumer
                rows += 1
        else:
            probecells = ws.cell(row=rows + 1, column=fromcolumns)
            while True:
                fromcells = ws.cell(row=rows, column=fromcolumns)
                tocells = ws.cell(row=rows, column=tocolumns)
                sumer = 0
                print(fromcells.value)
                for i in range(len(datalist)):
                    if datalist[i] in fromcells.value:
                        sumer += weightlist[i]
                tocells.value = sumer
                rows += 1
                probecells = ws.cell(row=rows, column=fromcolumns)
                if probecells.value is None:
                    break
        wb.save(self._excel_name)
        wb.close()

    """
    将加权后数据（dice）写入某一行中
    numbers为需要写入单元格数量，默认为读取内容遇到的第一个空单元格前同列某行
    columns默认为从第一行开始读入
    """

    def count_weight_sum_rows(self, data_weight_dice, fromrows, torows, numbers=None, columns=1):
        wb = load_workbook(self._path)
        ws = wb[self._sheet_name]
        datalist = list(data_weight_dice.keys())
        weightlist = list(data_weight_dice.values())
        fromcells = ws.cell(row=fromrows, column=columns)
        tocells = ws.cell(row=torows, column=columns)
        if numbers:
            for n in range(numbers):
                sumer = 0
                print('fromcells is ', fromcells.value)
                for i in range(len(datalist)):
                    if datalist[i] in fromcells.value:
                        print(fromcells.value)
                        sumer += weightlist[i]
                tocells.value = sumer
                columns += 1
        else:
            probecells = ws.cell(row=fromrows, column=columns + 1)
            while True:
                fromcells = ws.cell(row=fromrows, column=columns)
                tocells = ws.cell(row=torows, column=columns)
                sumer = 0
                print(fromcells.value)
                for i in range(len(datalist)):
                    if datalist[i] in fromcells.value:
                        sumer += weightlist[i]
                tocells.value = sumer
                columns += 1
                probecells = ws.cell(row=fromrows, column=columns)
                if probecells.value is None:
                    break
        wb.save(self._excel_name)
        wb.close()






if __name__ == '__main__':
    """
    请在初次使用时定义好下列变量：
    path:需要读取表格的绝对路径
    sheetname:需要操作的工作表名称
    请调用一次set_name_auto或是set_name定义保存名称（不加设置会保存在与脚本同目录下）
    其他可选方法：
    cat_columns
    cat_rows
    set_data_weight
    print_info
    print_weight_info
    save_to_cell_simple
    save_to_cell_rows
    save_to_cell_columns
    count_weight_sum_columns
    count_weight_sum_rows
    """

    """
    请务必填好下面内容
    path
    excelname
    sheetname
    """

    path = ''  # 要读取的表格的绝对路径
    excelnames = ''  # 选填，为保存名称
    sheetname = ''  # 需要操作的工作表名称


    excel = Excel(path, sheetname)
    if excelnames:
        excel.set_name(excelnames)
    else:
        excel.set_name_auto

    """
    此处还尚未写完，各位读者可凭自己需求修改
    未来哪天作者有空闲时则会继续完善
    """


